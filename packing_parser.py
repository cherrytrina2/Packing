from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side


SUPPLIER_OPTIONS: Dict[str, str] = {
    "自动识别": "auto",
    "BOE韩国模板(FP24PI004)": "boe_fp24pi004",
    "通用模板": "generic",
}

FIXED_TEMPLATE_PATH = Path(r"C:/Users/HP ELITEBOOK 640G10/Desktop/DONE/装箱方案公式模板.xlsx")
OUTPUT_FONT = Font(name="等线", size=12)
OUTPUT_ALIGN = Alignment(horizontal="center", vertical="center")
OUTPUT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def get_fixed_template_path() -> Path:
    if FIXED_TEMPLATE_PATH.exists():
        return FIXED_TEMPLATE_PATH
    fallback = Path.cwd() / "装箱方案公式模板.xlsx"
    return fallback

FIELD_ALIASES: Dict[str, List[str]] = {
    "seq_no": ["序号", "项次", "itemno", "index", "line", "no"],
    "crate_no": ["木箱号", "箱号", "cartonno", "case", "crate", "pallet", "pkg", "package", "colli"],
    "length_mm": ["长", "length", "len", "lmm", "l"],
    "width_mm": ["宽", "width", "wid", "wmm", "w"],
    "height_mm": ["高", "height", "ht", "hmm", "h"],
    "gross_weight_kg": ["毛重", "grossweight", "gwkgs", "gw", "weightkg", "gross"],
    "volume_cbm": ["体积", "volume", "cbm", "m3", "m^3"],
    "cargo_name": ["货物名称", "品名", "description", "cargo", "itemname", "name"],
    "qty": ["数量", "qty", "quantity", "q'ty", "qtty"],
}


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip().lower()
    s = s.replace("（", "(").replace("）", ")")
    s = re.sub(r"[\s_\-:/\\\[\]\(\)\.]+", "", s)
    return s


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", "")
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _header_score(row_vals: List[object]) -> Tuple[int, Dict[str, int]]:
    mapping: Dict[str, int] = {}
    score = 0
    normalized = [_norm(x) for x in row_vals]
    for field, aliases in FIELD_ALIASES.items():
        best_col = -1
        best_hit = ""
        for idx, val in enumerate(normalized):
            if not val:
                continue
            for a in aliases:
                aa = _norm(a)
                # avoid broad fuzzy hits; require exact or strong prefix for short aliases
                if aa and (aa == val or (len(aa) >= 4 and aa in val)):
                    if len(aa) > len(best_hit):
                        best_hit = aa
                        best_col = idx
        if best_col >= 0:
            mapping[field] = best_col
            score += 1
    return score, mapping


def detect_header(ws, scan_rows: int = 120) -> Tuple[int, Dict[str, int]]:
    best_row = -1
    best_map: Dict[str, int] = {}
    best_score = -1
    max_col = ws.max_column
    for r in range(1, min(ws.max_row, scan_rows) + 1):
        row_vals = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        score, mapping = _header_score(row_vals)
        if score > best_score:
            best_score = score
            best_row = r
            best_map = mapping
    if best_score < 3:
        raise ValueError("未识别到可用表头（至少需要识别3个关键字段）。")
    return best_row, best_map


def _sheet_ascii_fingerprint(ws) -> str:
    vals = []
    for r in range(1, min(ws.max_row, 12) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                vals.append(str(v).lower())
    return " ".join(vals)


def _detect_supplier(input_path: Path, ws, supplier: str) -> str:
    if supplier != "auto":
        return supplier
    fp = _sheet_ascii_fingerprint(ws) + " " + input_path.name.lower()
    if "fp24pi004" in fp or ("carton no" in fp and "packing list" in fp and "gw(kgs)" in fp):
        return "boe_fp24pi004"
    return "generic"


def _build_bound_groups(
    ws, data_start: int, l_col: int, w_col: int, h_col: int
) -> Tuple[Dict[int, List[int]], set[int], Dict[int, int]]:
    # Bound-unit rule: L/W/H columns are merged with the same row span.
    if l_col <= 0 or w_col <= 0 or h_col <= 0:
        return {}, set(), {}
    merge_l: Dict[int, Tuple[int, int]] = {}
    merge_w: Dict[int, Tuple[int, int]] = {}
    merge_h: Dict[int, Tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row < data_start:
            continue
        if mr.min_col == mr.max_col == l_col:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_l[r] = (mr.min_row, mr.max_row)
        if mr.min_col == mr.max_col == w_col:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_w[r] = (mr.min_row, mr.max_row)
        if mr.min_col == mr.max_col == h_col:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_h[r] = (mr.min_row, mr.max_row)

    starts: Dict[int, List[int]] = {}
    for r in range(data_start, ws.max_row + 1):
        a = merge_l.get(r)
        b = merge_w.get(r)
        c = merge_h.get(r)
        if a and b and c and a == b == c and a[1] > a[0]:
            starts.setdefault(a[0], list(range(a[0], a[1] + 1)))
    rows_in_group = {rr for rows in starts.values() for rr in rows}
    row_to_start: Dict[int, int] = {}
    for s, rows in starts.items():
        for rr in rows:
            row_to_start[rr] = s
    return starts, rows_in_group, row_to_start


def _is_total_row(seq, crate, cargo_name) -> bool:
    text = f"{seq} {crate} {cargo_name}".lower()
    text = text.replace(" ", "")
    keywords = ["total", "subtotal", "grandtotal", "合计", "总计", "小计", "합계"]
    return any(k in text for k in keywords)


def _parse_boe_fp24pi004(ws) -> pd.DataFrame:
    # Header occupies two rows; data starts from row 6 in this supplier template.
    start_row = 6
    group_start_to_rows, rows_in_group, row_to_start = _build_bound_groups(ws, start_row, l_col=11, w_col=12, h_col=13)
    records = []
    empty_streak = 0
    for r in range(start_row, ws.max_row + 1):
        group_start = row_to_start.get(r, r)
        group_rows = group_start_to_rows.get(group_start, [r])
        top = group_start
        seq = ws.cell(r, 2).value
        crate = ws.cell(r, 3).value
        cargo_name = ws.cell(r, 4).value
        qty = _to_float(ws.cell(r, 15).value)
        if qty is None:
            # Some supplier files only keep line quantity in QT'Y column.
            qty = _to_float(ws.cell(r, 5).value)
        gw = _to_float(ws.cell(r, 10).value)
        l = _to_float(ws.cell(r, 11).value)
        w = _to_float(ws.cell(r, 12).value)
        h = _to_float(ws.cell(r, 13).value)
        vol = _to_float(ws.cell(r, 14).value)
        # group member rows may have A/B empty due to merged cells, so read anchor row
        seq_anchor = ws.cell(top, 2).value
        crate_anchor = ws.cell(top, 3).value
        qty_anchor = _to_float(ws.cell(top, 15).value)
        gw_anchor = _to_float(ws.cell(top, 10).value)
        l_anchor = _to_float(ws.cell(top, 11).value)
        w_anchor = _to_float(ws.cell(top, 12).value)
        h_anchor = _to_float(ws.cell(top, 13).value)
        vol_anchor = _to_float(ws.cell(top, 14).value)
        if qty is None:
            qty = qty_anchor

        if all(v in (None, "") for v in (seq_anchor, crate_anchor, cargo_name, qty, gw_anchor, l_anchor, w_anchor, h_anchor, vol_anchor)):
            empty_streak += 1
            if empty_streak >= 20:
                break
            continue
        empty_streak = 0
        if _is_total_row(seq_anchor, crate_anchor, cargo_name):
            continue
        records.append(
            {
                "source_row": r,
                "seq_no": seq_anchor,
                "crate_no": crate_anchor,
                "cargo_name": cargo_name,
                "qty": qty,
                "length_mm": l_anchor,
                "width_mm": w_anchor,
                "height_mm": h_anchor,
                "gross_weight_kg": gw_anchor,
                "volume_cbm": vol_anchor,
                "merge_group_id": group_start,
                "merge_span": len(group_rows),
                "group_start": 1 if r == group_start else 0,
                "bound_rows": ",".join(str(x) for x in group_rows),
            }
        )
    if not records:
        raise ValueError("BOE模板解析失败：未识别到数据行。")
    return pd.DataFrame(records)


def parse_packing_list(input_path: Path, output_path: Path | None = None, supplier: str = "auto") -> Path:
    wb = load_workbook(input_path, data_only=True)
    ws = wb.active
    selected_supplier = _detect_supplier(input_path, ws, supplier)

    if selected_supplier == "boe_fp24pi004":
        df = _parse_boe_fp24pi004(ws)
        header_row = 5
        mapping = {
            "seq_no": 2,
            "crate_no": 3,
            "cargo_name": 4,
            "qty": 15,
            "gross_weight_kg": 10,
            "length_mm": 11,
            "width_mm": 12,
            "height_mm": 13,
            "volume_cbm": 14,
        }
    else:
        header_row, mapping = detect_header(ws)
        group_start_to_rows, rows_in_group, row_to_start = _build_bound_groups(
            ws,
            header_row + 1,
            l_col=(mapping.get("length_mm", -1) + 1) if "length_mm" in mapping else -1,
            w_col=(mapping.get("width_mm", -1) + 1) if "width_mm" in mapping else -1,
            h_col=(mapping.get("height_mm", -1) + 1) if "height_mm" in mapping else -1,
        )
        records = []
        empty_streak = 0
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            group_start = row_to_start.get(r, r)
            group_rows = group_start_to_rows.get(group_start, [r])
            anchor_row = [ws.cell(group_start, c).value for c in range(1, ws.max_column + 1)]
            if all(x in (None, "") for x in row):
                empty_streak += 1
                if empty_streak >= 20:
                    break
                continue
            empty_streak = 0

            def getv(field):
                ci = mapping.get(field)
                return row[ci] if ci is not None and ci < len(row) else None
            def geta(field):
                ci = mapping.get(field)
                return anchor_row[ci] if ci is not None and ci < len(anchor_row) else None

            rec = {
                "source_row": r,
                "seq_no": geta("seq_no"),
                "crate_no": geta("crate_no"),
                "cargo_name": getv("cargo_name"),
                "qty": (_to_float(getv("qty")) if _to_float(getv("qty")) is not None else _to_float(geta("qty"))),
                "length_mm": _to_float(geta("length_mm")),
                "width_mm": _to_float(geta("width_mm")),
                "height_mm": _to_float(geta("height_mm")),
                "gross_weight_kg": _to_float(geta("gross_weight_kg")),
                "volume_cbm": _to_float(geta("volume_cbm")),
                "merge_group_id": group_start,
                "merge_span": len(group_rows),
                "group_start": 1 if r == group_start else 0,
                "bound_rows": ",".join(str(x) for x in group_rows),
            }
            has_meaningful = any(
                rec[k] not in (None, "")
                for k in ["seq_no", "crate_no", "cargo_name", "qty", "length_mm", "width_mm", "height_mm", "gross_weight_kg"]
            )
            if has_meaningful and (not _is_total_row(rec.get("seq_no"), rec.get("crate_no"), rec.get("cargo_name"))):
                records.append(rec)
        if not records:
            raise ValueError("解析完成但未提取到数据行，请检查原始模板。")
        df = pd.DataFrame(records)

    if output_path is None:
        output_path = input_path.with_name(f"{input_path.stem}-parsed.xlsx")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="parsed_data", index=False)
        meta = pd.DataFrame(
            [
                {"key": "source_file", "value": str(input_path)},
                {"key": "sheet_name", "value": ws.title},
                {"key": "supplier_profile", "value": selected_supplier},
                {"key": "header_row", "value": header_row},
                {"key": "total_rows", "value": len(df)},
                {"key": "mapping", "value": str(mapping)},
            ]
        )
        meta.to_excel(writer, sheet_name="parse_meta", index=False)

    return output_path


def fill_template_from_parsed(parsed_path: Path, template_path: Path, output_path: Path | None = None) -> Path:
    if output_path is None:
        output_path = parsed_path.with_name(f"{parsed_path.stem}-template-filled.xlsx")

    df = pd.read_excel(parsed_path, sheet_name="parsed_data")
    wb = load_workbook(template_path)
    ws = wb.active

    data_start = 6
    clear_cols = {1, 2, 5, 6, 7, 8, 9, 10, 11, 12}
    to_unmerge = []
    for rg in ws.merged_cells.ranges:
        if rg.max_row < data_start:
            continue
        if any(c in clear_cols for c in range(rg.min_col, rg.max_col + 1)):
            to_unmerge.append(str(rg))
    for rg in to_unmerge:
        ws.unmerge_cells(rg)

    max_row = max(ws.max_row, data_start + len(df) + 20)
    for r in range(data_start, max_row + 1):
        for c in clear_cols:
            ws.cell(r, c).value = None

    row = data_start
    group_to_rows: Dict[str, List[int]] = {}
    for _, rec in df.iterrows():
        ws.cell(row, 1).value = rec.get("seq_no")
        ws.cell(row, 2).value = rec.get("crate_no")
        ws.cell(row, 3).value = rec.get("cargo_name")
        ws.cell(row, 4).value = rec.get("qty")
        ws.cell(row, 5).value = rec.get("length_mm")
        ws.cell(row, 6).value = rec.get("width_mm")
        ws.cell(row, 7).value = rec.get("height_mm")
        ws.cell(row, 8).value = rec.get("gross_weight_kg")
        ws.cell(row, 9).value = rec.get("volume_cbm")
        gid = str(rec.get("merge_group_id", f"r{row}"))
        group_to_rows.setdefault(gid, []).append(row)
        row += 1

    # keep A/B/E/F/G/H/I merged for indivisible units; C/D keep row-level display
    for rows in group_to_rows.values():
        if len(rows) <= 1:
            continue
        s = min(rows)
        e = max(rows)
        if e <= s:
            continue
        for c in (1, 2, 5, 6, 7, 8, 9):
            ws.merge_cells(start_row=s, start_column=c, end_row=e, end_column=c)

    end_row = row - 1
    for r in range(data_start, end_row + 1):
        for c in range(1, 13):
            cell = ws.cell(r, c)
            if isinstance(cell, MergedCell):
                continue
            cell.font = OUTPUT_FONT
            cell.alignment = OUTPUT_ALIGN
            cell.border = OUTPUT_BORDER

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Parse supplier packing list into normalized format")
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--supplier", choices=list(SUPPLIER_OPTIONS.values()), default="auto")
    parser.add_argument("--template", type=Path, default=None, help="填充目标模板路径")
    parser.add_argument("--template-output", type=Path, default=None)
    args = parser.parse_args()

    out = parse_packing_list(args.input, args.output, supplier=args.supplier)
    print(f"parsed output: {out}")
    if args.template:
        filled = fill_template_from_parsed(out, args.template, args.template_output)
        print(f"template output: {filled}")


if __name__ == "__main__":
    main()
