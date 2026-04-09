from __future__ import annotations

import argparse
import random
from dataclasses import dataclass, field
from copy import copy
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

DATA_START_ROW = 6
DATA_END_ROW = 995
CAR_INFO_COL = 10
CAR_INFO_SEQ_COL = 11
REMARK_COL = 12
OUTPUT_FONT = Font(name="等线", size=12)
OUTPUT_ALIGN = Alignment(horizontal="center", vertical="center")
OUTPUT_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

SIDE_WEIGHT_DIFF_LIMIT = 5000.0
BIG_PIECE_WEIGHT = 12000.0
LOW_LENGTH_UTIL_THRESHOLD = 0.90
UNDERFILLED_REPACK_MAX_ROUNDS = 3
UNDERFILLED_WIDTH_PENALTY_MULTIPLIER = 2.5
FR_WIDTH_REMARK_LIMITS = {"20FR": 2226.0, "40FR": 2374.0}
FR_HEIGHT_REMARK_LIMIT = 1900.0
MODEL_DISPLAY_NAMES = {"710板": "710定制板", "880板": "880定制板"}
PARALLEL_MODEL_LIMITS: Dict[str, Tuple[float, float]] = {
    "20GP": (2300.0, 2150.0),  # 2200 minus door reserve
    "40GP": (2300.0, 2200.0),
    "40HQ": (2300.0, 2500.0),
    "20FR": (2226.0, 4500.0),
    "40FR": (2374.0, 4500.0),
}


@dataclass
class Item:
    row: int
    length: float
    width: float
    height: float
    weight: float = 0.0
    bound_rows: List[int] = field(default_factory=list)


@dataclass
class Unit:
    items: List[Item] = field(default_factory=list)
    length: float = 0.0
    width: float = 0.0
    height: float = 0.0
    weight: float = 0.0

    @classmethod
    def from_item(cls, item: Item) -> "Unit":
        return cls(items=[item], length=item.length, width=item.width, height=item.height, weight=item.weight)

    def rows(self) -> List[int]:
        out: List[int] = []
        for i in self.items:
            out.extend(i.bound_rows if i.bound_rows else [i.row])
        return sorted(set(out))


@dataclass(frozen=True)
class BoxRule:
    code: str
    length_cap: float
    max_width: float | None
    max_height: float | None
    max_payload: float | None
    width_hard_limit: float
    width_penalty: float = 3.0
    length_fill_weight: float = 1.0


BOX_RULES: Dict[str, BoxRule] = {
    "20GP": BoxRule("20GP", 5750, 2300, 2200, 21670, 900),
    "40GP": BoxRule("40GP", 11450, 2300, 2200, 26480, 900),
    "40HQ": BoxRule("40HQ", 11450, 2300, 2500, 26480, 900),
    "20FR": BoxRule("20FR", 5450, None, 4500, 31200, 2500),
    "40FR": BoxRule("40FR", 11250, None, 4500, 40000, 2500),
    "710板": BoxRule("710板", 6700, 5800, 4500, 70000, 3500),
    "880板": BoxRule("880板", 8100, 5800, 4500, 80000, 3500),
}

BOARD_CODES = [c for c in BOX_RULES if "710" in c or "880" in c]

SCENARIO_LABELS: Dict[str, str] = {
    "scenario1": "??1(40HQ+FR)",
    "scenario2": "??2(FR-only)",
    "scenario3": "??3(???:GP+HQ+FR)",
    "auto": "????(????HQ)",
    "custom": "?????",
}


@dataclass
class Bin:
    units: List[Unit] = field(default_factory=list)
    used_length: float = 0.0
    used_weight: float = 0.0
    width_min: float = 0.0
    width_max: float = 0.0

    def can_fit(self, unit: Unit, length_cap: float, payload_cap: float | None) -> bool:
        if self.used_length + unit.length > length_cap:
            return False
        if payload_cap is not None and self.used_weight + unit.weight > payload_cap:
            return False
        return True

    def add(self, unit: Unit):
        self.units.append(unit)
        self.used_length += unit.length
        self.used_weight += unit.weight
        widths = [u.width for u in self.units]
        self.width_min = min(widths)
        self.width_max = max(widths)

    def remove(self, unit: Unit):
        self.units.remove(unit)
        self.used_length -= unit.length
        self.used_weight -= unit.weight
        if self.units:
            widths = [u.width for u in self.units]
            self.width_min = min(widths)
            self.width_max = max(widths)
        else:
            self.width_min = 0.0
            self.width_max = 0.0


def estimate_side_diff(units: List[Unit]) -> float:
    left = 0.0
    right = 0.0
    for u in sorted(units, key=lambda x: x.weight, reverse=True):
        if left <= right:
            left += u.weight
        else:
            right += u.weight
    return abs(left - right)


def load_items(excel_path: Path) -> List[Item]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active

    merge_l: Dict[int, Tuple[int, int]] = {}
    merge_w: Dict[int, Tuple[int, int]] = {}
    merge_h: Dict[int, Tuple[int, int]] = {}
    for mr in ws.merged_cells.ranges:
        if mr.min_row < DATA_START_ROW:
            continue
        if mr.min_col == mr.max_col == 5:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_l[r] = (mr.min_row, mr.max_row)
        if mr.min_col == mr.max_col == 6:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_w[r] = (mr.min_row, mr.max_row)
        if mr.min_col == mr.max_col == 7:
            for r in range(mr.min_row, mr.max_row + 1):
                merge_h[r] = (mr.min_row, mr.max_row)

    group_start_to_rows: Dict[int, List[int]] = {}
    max_row = max(ws.max_row, DATA_END_ROW)
    for r in range(DATA_START_ROW, max_row + 1):
        ml = merge_l.get(r)
        mw = merge_w.get(r)
        mh = merge_h.get(r)
        if ml and mw and mh and ml == mw == mh and ml[1] > ml[0]:
            group_start_to_rows.setdefault(ml[0], list(range(ml[0], ml[1] + 1)))

    rows_in_group = {rr for rows in group_start_to_rows.values() for rr in rows}
    starts = set(group_start_to_rows.keys())

    items: List[Item] = []
    for r in range(DATA_START_ROW, max_row + 1):
        if r in rows_in_group and r not in starts:
            continue
        bound_rows = group_start_to_rows.get(r, [r])
        l = ws.cell(row=r, column=5).value
        w = ws.cell(row=r, column=6).value
        h = ws.cell(row=r, column=7).value
        if pd.isna(l) or pd.isna(w) or pd.isna(h):
            continue
        weight = ws.cell(row=r, column=8).value
        items.append(
            Item(
                row=r,
                length=float(l),
                width=float(w),
                height=float(h),
                weight=float(weight) if (weight is not None and not pd.isna(weight)) else 0.0,
                bound_rows=bound_rows,
            )
        )
    return items


def build_parallel_groups(
    items: List[Item], width_limit: float = 2300.0, height_limit: float = 2500.0
) -> Tuple[List[Unit], List[Item]]:
    eligible = [i for i in items if i.width <= width_limit and i.height <= height_limit]
    leftovers = [i for i in items if i not in eligible]
    eligible.sort(key=lambda x: (x.width, x.length), reverse=True)
    if len(eligible) <= 1:
        return [Unit.from_item(i) for i in eligible], leftovers

    by_row = {i.row: i for i in eligible}
    partner: Dict[int, int | None] = {i.row: None for i in eligible}
    w_scale = max(i.width for i in eligible) or 1.0
    l_scale = max(i.length for i in eligible) or 1.0

    def pair_score(a: Item, b: Item) -> float:
        if a.row == b.row:
            return float("inf")
        pair_w = a.width + b.width
        if pair_w > width_limit:
            return float("inf")
        if abs(a.width - b.width) / w_scale > 0.55:
            return float("inf")
        if abs(a.length - b.length) / l_scale > 0.55:
            return float("inf")
        w_delta = abs(a.width - b.width) / w_scale
        l_delta = abs(a.length - b.length) / l_scale
        width_gap = (width_limit - pair_w) / width_limit
        return (w_delta * 0.55) + (l_delta * 0.35) + (width_gap * 0.10)

    def break_pair(r: int):
        p = partner[r]
        if p is None:
            return
        partner[r] = None
        partner[p] = None

    changed = True
    rounds = 0
    while changed and rounds < max(20, len(eligible) * 6):
        rounds += 1
        changed = False
        for a in eligible:
            best_b = None
            best = float("inf")
            for b in eligible:
                if b.row == a.row:
                    continue
                s = pair_score(a, b)
                if s < best:
                    best = s
                    best_b = b.row
            if best_b is None or best == float("inf"):
                continue
            cur = partner[a.row]
            cur_s = pair_score(a, by_row[cur]) if cur is not None else float("inf")
            if cur is not None and best >= cur_s * 0.99:
                continue
            b_cur = partner[best_b]
            b_cur_s = pair_score(by_row[best_b], by_row[b_cur]) if b_cur is not None else float("inf")
            if b_cur is not None and best >= b_cur_s * 0.99:
                continue
            break_pair(a.row)
            break_pair(best_b)
            partner[a.row] = best_b
            partner[best_b] = a.row
            changed = True

    units: List[Unit] = []
    used: set[int] = set()
    for a in eligible:
        if a.row in used:
            continue
        b_row = partner[a.row]
        if b_row is not None and b_row not in used:
            b = by_row[b_row]
            grp = [a, b]
            used.add(a.row)
            used.add(b_row)
        else:
            grp = [a]
            used.add(a.row)
        units.append(
            Unit(
                items=grp,
                length=max(i.length for i in grp),
                width=sum(i.width for i in grp),
                height=max(i.height for i in grp),
                weight=sum(i.weight for i in grp),
            )
        )
    return units, leftovers


def parallel_pair_score(
    a: Item, b: Item, w_scale: float, l_scale: float, width_limit: float = 2300.0, height_limit: float = 2500.0
) -> float:
    if a.row == b.row:
        return float("inf")
    if a.width > width_limit or b.width > width_limit:
        return float("inf")
    if a.height > height_limit or b.height > height_limit:
        return float("inf")
    pair_w = a.width + b.width
    if pair_w > width_limit:
        return float("inf")
    if abs(a.width - b.width) / (w_scale or 1.0) > 0.55:
        return float("inf")
    if abs(a.length - b.length) / (l_scale or 1.0) > 0.55:
        return float("inf")
    w_delta = abs(a.width - b.width) / (w_scale or 1.0)
    l_delta = abs(a.length - b.length) / (l_scale or 1.0)
    width_gap = (width_limit - pair_w) / width_limit
    return (w_delta * 0.55) + (l_delta * 0.35) + (width_gap * 0.10)


def infer_parallel_pairs_in_same_box(
    assignments: Dict[int, str], items: List[Item], model: str
) -> List[Tuple[Item, Item]]:
    limits = PARALLEL_MODEL_LIMITS.get(model)
    if limits is None:
        return []
    width_limit, height_limit = limits
    box_prefix = f"{model}-"
    item_box: Dict[int, str] = {}
    by_row = {it.row: it for it in items}
    for it in items:
        box = None
        for r in (it.bound_rows if it.bound_rows else [it.row]):
            if r in assignments:
                box = str(assignments[r])
                break
        if box is not None and box.startswith(box_prefix):
            item_box[it.row] = box

    boxes: Dict[str, List[Item]] = {}
    for row, box in item_box.items():
        boxes.setdefault(box, []).append(by_row[row])

    out_pairs: List[Tuple[Item, Item]] = []
    for _, box_items in boxes.items():
        if len(box_items) < 2:
            continue
        w_scale = max(i.width for i in box_items) or 1.0
        l_scale = max(i.length for i in box_items) or 1.0
        candidates: List[Tuple[float, int, int]] = []
        for i in range(len(box_items)):
            for j in range(i + 1, len(box_items)):
                a = box_items[i]
                b = box_items[j]
                s = parallel_pair_score(a, b, w_scale, l_scale, width_limit=width_limit, height_limit=height_limit)
                if s != float("inf"):
                    candidates.append((s, a.row, b.row))
        candidates.sort(key=lambda x: x[0])
        used: set[int] = set()
        for _, ra, rb in candidates:
            if ra in used or rb in used:
                continue
            used.add(ra)
            used.add(rb)
            out_pairs.append((by_row[ra], by_row[rb]))
    return out_pairs


def infer_parallel_merge_rows_from_assignments(
    assignments: Dict[int, str], items: List[Item], models: List[str] | None = None
) -> List[List[int]]:
    merges: List[List[int]] = []
    target_models = models or list(PARALLEL_MODEL_LIMITS.keys())
    for model in target_models:
        for a, b in infer_parallel_pairs_in_same_box(assignments, items, model=model):
            rows = sorted(set((a.bound_rows if a.bound_rows else [a.row]) + (b.bound_rows if b.bound_rows else [b.row])))
            if len(rows) > 1:
                merges.append(rows)
    return merges


def repack_model_with_inferred_parallel(assignments: Dict[int, str], items: List[Item], model: str) -> Dict[int, str]:
    if model not in PARALLEL_MODEL_LIMITS or model not in BOX_RULES:
        return assignments
    pairs = infer_parallel_pairs_in_same_box(assignments, items, model=model)
    if not pairs:
        return assignments

    model_items: List[Item] = []
    for it in items:
        box = None
        for r in (it.bound_rows if it.bound_rows else [it.row]):
            if r in assignments:
                box = str(assignments[r])
                break
        if box is not None and box.startswith(f"{model}-"):
            model_items.append(it)
    if len(model_items) <= 1:
        return assignments

    pair_map: Dict[int, int] = {}
    by_row = {it.row: it for it in model_items}
    for a, b in pairs:
        if a.row in by_row and b.row in by_row:
            pair_map[a.row] = b.row
            pair_map[b.row] = a.row

    units: List[Unit] = []
    used: set[int] = set()
    for it in model_items:
        if it.row in used:
            continue
        p = pair_map.get(it.row)
        if p is not None and p in by_row and p not in used:
            b = by_row[p]
            grp = [it, b]
            used.add(it.row)
            used.add(p)
            units.append(
                Unit(
                    items=grp,
                    length=max(x.length for x in grp),
                    width=sum(x.width for x in grp),
                    height=max(x.height for x in grp),
                    weight=sum(x.weight for x in grp),
                )
            )
        else:
            used.add(it.row)
            units.append(Unit.from_item(it))

    rule = BOX_RULES[model]
    bins = place_units(
        units,
        rule.length_cap,
        rule.max_payload,
        rule.width_hard_limit,
        rule.width_penalty,
        rule.length_fill_weight,
        order_mode="length",
        random_seed=0,
    )
    bins = improve_bins(bins, rule.length_cap, rule.max_payload, rule.width_hard_limit)
    repacked = bins_to_assignment(bins, f"{model}-")

    old_boxes = {str(v) for v in assignments.values() if str(v).startswith(f"{model}-")}
    new_boxes = {str(v) for v in repacked.values() if str(v).startswith(f"{model}-")}
    if not repacked or len(new_boxes) > len(old_boxes):
        return assignments

    out = assignments.copy()
    model_rows = {r for it in model_items for r in (it.bound_rows if it.bound_rows else [it.row])}
    for r in model_rows:
        out.pop(r, None)
    out.update(repacked)
    return out


def _box_seq_key(box_name: str) -> Tuple[int, str]:
    try:
        suffix = box_name.split("-", 1)[1]
    except Exception:
        return (10**9, box_name)
    digits = ""
    for ch in suffix:
        if ch.isdigit():
            digits += ch
        else:
            break
    if digits:
        return (int(digits), box_name)
    return (10**9, box_name)


def _model_units_by_box(assignments: Dict[int, str], items: List[Item], model: str) -> Dict[str, List[Unit]]:
    pairs = infer_parallel_pairs_in_same_box(assignments, items, model=model)
    item_by_row = {it.row: it for it in items}
    pair_map: Dict[int, int] = {}
    for a, b in pairs:
        pair_map[a.row] = b.row
        pair_map[b.row] = a.row

    box_to_items: Dict[str, List[Item]] = {}
    for it in items:
        box = None
        for r in (it.bound_rows if it.bound_rows else [it.row]):
            if r in assignments:
                box = str(assignments[r])
                break
        if box is not None and box.startswith(f"{model}-"):
            box_to_items.setdefault(box, []).append(it)

    out: Dict[str, List[Unit]] = {}
    for box, its in box_to_items.items():
        used: set[int] = set()
        units: List[Unit] = []
        by_row = {it.row: it for it in its}
        for it in its:
            if it.row in used:
                continue
            p = pair_map.get(it.row)
            if p is not None and p in by_row and p not in used:
                jt = by_row[p]
                grp = [it, jt]
                used.add(it.row)
                used.add(jt.row)
                units.append(
                    Unit(
                        items=grp,
                        length=max(x.length for x in grp),
                        width=sum(x.width for x in grp),
                        height=max(x.height for x in grp),
                        weight=sum(x.weight for x in grp),
                    )
                )
            else:
                used.add(it.row)
                units.append(Unit.from_item(it))
        out[box] = units
    return out


def fill_small_units_between_boxes(assignments: Dict[int, str], items: List[Item], model: str) -> Dict[int, str]:
    if model not in BOX_RULES:
        return assignments
    rule = BOX_RULES[model]
    box_units = _model_units_by_box(assignments, items, model=model)
    if len(box_units) <= 1:
        return assignments

    changed = True
    while changed:
        changed = False
        box_names = sorted(box_units.keys(), key=_box_seq_key)
        for dst_name in box_names:
            if dst_name not in box_units:
                continue
            dst_units = box_units[dst_name]
            dst_len = used_length_of_units(dst_units)
            dst_wt = sum(u.weight for u in dst_units)
            dst_widths = [u.width for u in dst_units]
            dst_wmin = min(dst_widths) if dst_widths else 0.0
            dst_wmax = max(dst_widths) if dst_widths else 0.0

            # Fill earlier boxes first: try pulling short units from later boxes.
            later_boxes = [b for b in box_names if _box_seq_key(b)[0] > _box_seq_key(dst_name)[0] and b in box_units]
            for src_name in reversed(later_boxes):
                src_units = box_units.get(src_name, [])
                if not src_units:
                    continue
                # Prefer short/light units to maximize fit opportunities.
                candidates = sorted(src_units, key=lambda u: (u.length, u.weight))
                picked_idx = -1
                for idx, u in enumerate(candidates):
                    if dst_len + u.length > rule.length_cap:
                        continue
                    if rule.max_payload is not None and dst_wt + u.weight > rule.max_payload:
                        continue
                    nwmin = min(dst_wmin, u.width) if dst_units else u.width
                    nwmax = max(dst_wmax, u.width) if dst_units else u.width
                    if nwmax - nwmin > rule.width_hard_limit:
                        continue
                    cand_units = dst_units + [u]
                    side_diff = estimate_side_diff(cand_units)
                    if any(x.weight >= BIG_PIECE_WEIGHT for x in cand_units) and side_diff > SIDE_WEIGHT_DIFF_LIMIT:
                        continue
                    picked_idx = idx
                    break
                if picked_idx < 0:
                    continue

                u = candidates[picked_idx]
                # Remove by identity from source.
                for i, uu in enumerate(src_units):
                    if uu is u:
                        src_units.pop(i)
                        break
                dst_units.append(u)
                changed = True
                if not src_units:
                    del box_units[src_name]
                break
            if changed:
                break

    out = assignments.copy()
    # Clear model rows then re-assign from updated box units.
    model_rows: set[int] = set()
    for it in items:
        for r in (it.bound_rows if it.bound_rows else [it.row]):
            b = assignments.get(r)
            if b is not None and str(b).startswith(f"{model}-"):
                model_rows.add(r)
    for r in model_rows:
        out.pop(r, None)
    out.update(assignments_from_box_units(box_units))
    return out


def apply_parallel_repack_and_merge(
    assignments: Dict[int, str], items: List[Item], allowed_codes: List[str] | None = None
) -> Tuple[Dict[int, str], List[List[int]]]:
    target_models = [m for m in (allowed_codes or list(BOX_RULES.keys())) if m in PARALLEL_MODEL_LIMITS]
    out = assignments.copy()
    for model in target_models:
        out = repack_model_with_inferred_parallel(out, items, model=model)
        out = fill_small_units_between_boxes(out, items, model=model)
    merges = infer_parallel_merge_rows_from_assignments(out, items, models=target_models)
    return out, merges


def place_units(
    units: List[Unit],
    length_cap: float,
    payload_cap: float | None,
    width_hard_limit: float,
    width_penalty: float,
    length_fill_weight: float,
    order_mode: str = "length",
    random_seed: int = 0,
) -> List[Bin]:
    if order_mode == "as_is":
        ordered = units[:]
    elif order_mode == "weight":
        ordered = sorted(units, key=lambda u: (u.weight, u.length), reverse=True)
    elif order_mode == "width":
        ordered = sorted(units, key=lambda u: (u.width, u.length), reverse=True)
    elif order_mode == "hybrid":
        ordered = sorted(units, key=lambda u: (u.length * 0.6 + u.weight * 0.4, u.width), reverse=True)
    elif order_mode == "random":
        ordered = units[:]
        random.Random(random_seed).shuffle(ordered)
    else:
        ordered = sorted(units, key=lambda u: (u.length, u.width), reverse=True)

    bins: List[Bin] = []
    for unit in ordered:
        best_idx = -1
        best_score = float("inf")
        for i, b in enumerate(bins):
            if not b.can_fit(unit, length_cap, payload_cap):
                continue
            nwmin = min(b.width_min, unit.width) if b.units else unit.width
            nwmax = max(b.width_max, unit.width) if b.units else unit.width
            span = nwmax - nwmin
            if span > width_hard_limit:
                continue
            left = length_cap - (b.used_length + unit.length)
            candidate_units = b.units + [unit]
            side_diff = estimate_side_diff(candidate_units)
            has_big = any(u.weight >= BIG_PIECE_WEIGHT for u in candidate_units)
            balance_penalty = side_diff * 0.05
            if has_big and side_diff > SIDE_WEIGHT_DIFF_LIMIT:
                balance_penalty += (side_diff - SIDE_WEIGHT_DIFF_LIMIT) * 10.0
            w_left = (payload_cap - (b.used_weight + unit.weight)) if payload_cap is not None else 0.0
            score = span * width_penalty + left * length_fill_weight + max(0.0, w_left) * 0.02 + balance_penalty
            if score < best_score:
                best_score = score
                best_idx = i
        if best_idx >= 0:
            bins[best_idx].add(unit)
        else:
            nb = Bin()
            nb.add(unit)
            bins.append(nb)
    return bins


def improve_bins(bins: List[Bin], length_cap: float, payload_cap: float | None, width_hard_limit: float) -> List[Bin]:
    improved = True
    while improved:
        improved = False
        bins.sort(key=lambda b: b.used_length)
        for i in range(len(bins)):
            src = bins[i]
            src_units = sorted(src.units[:], key=lambda u: u.length, reverse=True)
            temp_len = [b.used_length for b in bins]
            temp_wgt = [b.used_weight for b in bins]
            temp_min = [b.width_min for b in bins]
            temp_max = [b.width_max for b in bins]
            temp_has = [bool(b.units) for b in bins]
            plan: List[Tuple[Unit, int]] = []
            can_close = True

            for unit in src_units:
                pick = -1
                best_left = float("inf")
                for j in range(len(bins)):
                    if j == i:
                        continue
                    if temp_len[j] + unit.length > length_cap:
                        continue
                    if payload_cap is not None and temp_wgt[j] + unit.weight > payload_cap:
                        continue
                    if temp_has[j]:
                        nwmin = min(temp_min[j], unit.width)
                        nwmax = max(temp_max[j], unit.width)
                    else:
                        nwmin = unit.width
                        nwmax = unit.width
                    if nwmax - nwmin > width_hard_limit:
                        continue
                    cand_units = bins[j].units + [unit]
                    side_diff = estimate_side_diff(cand_units)
                    if any(x.weight >= BIG_PIECE_WEIGHT for x in cand_units) and side_diff > SIDE_WEIGHT_DIFF_LIMIT:
                        continue
                    left = length_cap - (temp_len[j] + unit.length)
                    if left < best_left:
                        best_left = left
                        pick = j
                if pick < 0:
                    can_close = False
                    break
                plan.append((unit, pick))
                temp_len[pick] += unit.length
                temp_wgt[pick] += unit.weight
                if temp_has[pick]:
                    temp_min[pick] = min(temp_min[pick], unit.width)
                    temp_max[pick] = max(temp_max[pick], unit.width)
                else:
                    temp_min[pick] = unit.width
                    temp_max[pick] = unit.width
                    temp_has[pick] = True

            if can_close:
                for unit, pick in plan:
                    src.remove(unit)
                    bins[pick].add(unit)
                bins.pop(i)
                improved = True
                break
    return bins


def unit_fits_rule(unit: Unit, rule: BoxRule) -> bool:
    if unit.length > rule.length_cap:
        return False
    if rule.max_width is not None and unit.width > rule.max_width:
        return False
    max_h = rule.max_height
    # 20GP door-height reserve: keep 50mm margin.
    if rule.code == "20GP" and max_h is not None:
        max_h = max_h - 50
    if max_h is not None and unit.height > max_h:
        return False
    if rule.max_payload is not None and unit.weight > rule.max_payload:
        return False
    return True


def fr_volume_utilization(unit: Unit, code: str) -> float:
    """
    Estimate FR volume utilization by practical effective width.
    """
    if code not in ("20FR", "40FR"):
        return -1.0
    rule = BOX_RULES[code]
    eff_width = 2226.0 if code == "20FR" else 2374.0
    max_h = rule.max_height or 4500.0
    cap = rule.length_cap * eff_width * max_h
    if cap <= 0:
        return -1.0
    vol = max(0.0, unit.length) * max(0.0, unit.width) * max(0.0, unit.height)
    return vol / cap


def pick_fr_by_volume_utilization(unit: Unit, allowed_codes: List[str]) -> str | None:
    fit_20 = "20FR" in allowed_codes and unit_fits_rule(unit, BOX_RULES["20FR"])
    fit_40 = "40FR" in allowed_codes and unit_fits_rule(unit, BOX_RULES["40FR"])
    if fit_20 and fit_40:
        return "20FR" if fr_volume_utilization(unit, "20FR") >= fr_volume_utilization(unit, "40FR") else "40FR"
    if fit_20:
        return "20FR"
    if fit_40:
        return "40FR"
    return None


def unit_fits_any_standard_fr_gp(unit: Unit) -> bool:
    return any(unit_fits_rule(unit, BOX_RULES[c]) for c in ("20GP", "40GP", "20FR", "40FR"))


def unit_can_use_custom_board(unit: Unit) -> bool:
    return any(i.weight > 40000 for i in unit.items) and (not unit_fits_any_standard_fr_gp(unit))


def bins_to_assignment(bins: List[Bin], prefix: str) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for idx, b in enumerate(bins, start=1):
        box = f"{prefix}{idx}"
        for u in b.units:
            for r in u.rows():
                out[r] = box
    return out


def pack_by_rule_priority_mode(
    units: List[Unit], rule_codes: List[str], order_mode: str = "length", random_seed: int = 0
) -> Dict[int, str]:
    allowed_codes = [c for c in rule_codes if c in BOX_RULES]
    remaining = units[:]
    out: Dict[int, str] = {}
    for code in allowed_codes:
        rule = BOX_RULES[code]
        eligible = [u for u in remaining if unit_fits_rule(u, rule)]
        if code != "40HQ":
            eligible = [u for u in eligible if len(u.items) == 1]
        if code in BOARD_CODES:
            eligible = [u for u in eligible if unit_can_use_custom_board(u)]
        if not eligible:
            continue
        bins = place_units(
            eligible,
            rule.length_cap,
            rule.max_payload,
            rule.width_hard_limit,
            rule.width_penalty,
            rule.length_fill_weight,
            order_mode=order_mode,
            random_seed=random_seed,
        )
        bins = improve_bins(bins, rule.length_cap, rule.max_payload, rule.width_hard_limit)
        partial = bins_to_assignment(bins, f"{code}-")
        out.update(partial)
        assigned = set(partial.keys())
        remaining = [u for u in remaining if all(r not in assigned for r in u.rows())]

    fallback_idx = 1
    # Fallback (last few leftovers) uses small-box-first priority when available.
    fallback_priority = ["20GP", "20FR", "40FR", "40HQ"]
    ordered_fallback = [c for c in fallback_priority if c in allowed_codes]
    for c in allowed_codes:
        if c not in ordered_fallback and c in BOX_RULES:
            ordered_fallback.append(c)

    for u in remaining:
        picked = None
        fr_pick = pick_fr_by_volume_utilization(u, ordered_fallback)
        if fr_pick is not None:
            picked = fr_pick
        for code in ordered_fallback:
            if picked is not None:
                break
            if code in BOARD_CODES and (not unit_can_use_custom_board(u)):
                continue
            if unit_fits_rule(u, BOX_RULES[code]):
                picked = code
                break
        if picked is None and ordered_fallback:
            picked = ordered_fallback[0]
        if picked is None:
            picked = "40FR"
        for r in u.rows():
            out[r] = f"{picked}-FALLBACK{fallback_idx}"
        fallback_idx += 1
    return out


def assignment_uses_model(assignments: Dict[int, str], model: str) -> bool:
    prefix = f"{model}-"
    return any(str(v).startswith(prefix) for v in assignments.values())


def used_length_of_units(units: List[Unit]) -> float:
    """
    Length utilization uses occupied longitudinal length.
    For parallel-packed cargo, Unit.length already stores max(item.length),
    so summing Unit.length applies "parallel takes the longer length" rule.
    """
    return sum(u.length for u in units)


def box_units_from_assignments(assignments: Dict[int, str], units: List[Unit]) -> Dict[str, List[Unit]]:
    """
    Build box -> units map from assignments.
    Each unit is counted once by the first assigned row it contains.
    """
    box_units: Dict[str, List[Unit]] = {}
    for u in units:
        box_name = None
        for r in u.rows():
            if r in assignments:
                box_name = str(assignments[r])
                break
        if box_name is None:
            continue
        box_units.setdefault(box_name, []).append(u)
    return box_units


def has_underfilled_boxes(
    assignments: Dict[int, str], units: List[Unit], length_util_threshold: float = LOW_LENGTH_UTIL_THRESHOLD
) -> bool:
    box_units = box_units_from_assignments(assignments, units)
    for box_name, us in box_units.items():
        model = box_name.split("-", 1)[0]
        rule = BOX_RULES.get(model)
        if rule is None or rule.length_cap <= 0:
            continue
        used_length = used_length_of_units(us)
        util = used_length / rule.length_cap
        if util < length_util_threshold:
            return True
    return False


def underfilled_rows(
    assignments: Dict[int, str], units: List[Unit], length_util_threshold: float = LOW_LENGTH_UTIL_THRESHOLD
) -> set[int]:
    out: set[int] = set()
    box_units = box_units_from_assignments(assignments, units)
    for box_name, us in box_units.items():
        model = box_name.split("-", 1)[0]
        rule = BOX_RULES.get(model)
        if rule is None or rule.length_cap <= 0:
            continue
        used_length = used_length_of_units(us)
        util = used_length / rule.length_cap
        if util < length_util_threshold:
            for u in us:
                out.update(u.rows())
    return out


def assignments_from_box_units(box_units: Dict[str, List[Unit]]) -> Dict[int, str]:
    out: Dict[int, str] = {}
    for box_name, us in box_units.items():
        for u in us:
            for r in u.rows():
                out[r] = box_name
    return out


def consolidate_tiny_underfilled_boxes(assignments: Dict[int, str], units: List[Unit]) -> Dict[int, str]:
    """
    Post-process: merge tiny underfilled boxes into other opened boxes when feasible.
    This targets pathological low-util boxes (like a tiny single piece in one box).
    Width similarity is still preferred when picking target boxes.
    """
    box_units = box_units_from_assignments(assignments, units)
    if not box_units:
        return assignments

    changed = True
    while changed:
        changed = False
        # Candidate source boxes: low length utilization and very short used length.
        candidates: List[Tuple[str, float]] = []
        for box_name, us in box_units.items():
            model = box_name.split("-", 1)[0]
            rule = BOX_RULES.get(model)
            if rule is None or rule.length_cap <= 0:
                continue
            used_len = used_length_of_units(us)
            util = used_len / rule.length_cap
            if util < LOW_LENGTH_UTIL_THRESHOLD and used_len <= rule.length_cap * 0.45:
                candidates.append((box_name, util))
        candidates.sort(key=lambda x: x[1])

        for src_name, _ in candidates:
            if src_name not in box_units:
                continue
            src_units = box_units[src_name]
            src_model = src_name.split("-", 1)[0]
            src_rule = BOX_RULES.get(src_model)
            if src_rule is None:
                continue
            src_len = used_length_of_units(src_units)
            src_wt = sum(u.weight for u in src_units)
            src_widths = [u.width for u in src_units]
            src_wmin = min(src_widths) if src_widths else 0.0
            src_wmax = max(src_widths) if src_widths else 0.0

            best_target = None
            best_score = float("inf")
            for dst_name, dst_units in box_units.items():
                if dst_name == src_name:
                    continue
                dst_model = dst_name.split("-", 1)[0]
                rule = BOX_RULES.get(dst_model)
                if rule is None:
                    continue
                if any(not unit_fits_rule(u, rule) for u in src_units):
                    continue

                dst_len = used_length_of_units(dst_units)
                if dst_len + src_len > rule.length_cap:
                    continue
                dst_wt = sum(u.weight for u in dst_units)
                if rule.max_payload is not None and dst_wt + src_wt > rule.max_payload:
                    continue

                dst_widths = [u.width for u in dst_units]
                dst_wmin = min(dst_widths) if dst_widths else 0.0
                dst_wmax = max(dst_widths) if dst_widths else 0.0
                nwmin = min(dst_wmin, src_wmin)
                nwmax = max(dst_wmax, src_wmax)
                span = nwmax - nwmin
                # Try strict first; for tiny/light sources allow one relaxed fallback.
                span_limit_strict = rule.width_hard_limit * 1.25
                span_limit_relaxed = rule.width_hard_limit * 2.20
                if span > span_limit_strict:
                    tiny_light_src = (src_len <= src_rule.length_cap * 0.40) and (src_wt <= 2000.0)
                    if (not tiny_light_src) or span > span_limit_relaxed:
                        continue

                merged_units = dst_units + src_units
                side_diff = estimate_side_diff(merged_units)
                if any(u.weight >= BIG_PIECE_WEIGHT for u in merged_units) and side_diff > SIDE_WEIGHT_DIFF_LIMIT:
                    continue

                # Width similarity first, then better length fill.
                left = rule.length_cap - (dst_len + src_len)
                score = span * (rule.width_penalty * 3.0) + left
                if score < best_score:
                    best_score = score
                    best_target = dst_name

            if best_target is not None:
                box_units[best_target].extend(src_units)
                del box_units[src_name]
                changed = True
                break

    return assignments_from_box_units(box_units)


def repack_prioritize_merging_underfilled(
    units: List[Unit], rule_codes: List[str], base_assignments: Dict[int, str]
) -> Dict[int, str]:
    """
    Re-pack once by delaying units from underfilled boxes, so they are
    attempted to be merged into already-opened boxes first.
    """
    marked_rows = underfilled_rows(base_assignments, units)
    if not marked_rows:
        return base_assignments.copy()

    allowed_codes = [c for c in rule_codes if c in BOX_RULES]
    remaining = units[:]
    out: Dict[int, str] = {}

    for code in allowed_codes:
        rule = BOX_RULES[code]
        eligible = [u for u in remaining if unit_fits_rule(u, rule)]
        if code != "40HQ":
            eligible = [u for u in eligible if len(u.items) == 1]
        if code in BOARD_CODES:
            eligible = [u for u in eligible if unit_can_use_custom_board(u)]
        if not eligible:
            continue

        normal = sorted(
            [u for u in eligible if not any(r in marked_rows for r in u.rows())],
            key=lambda u: (u.length, u.width),
            reverse=True,
        )
        underfilled = sorted(
            [u for u in eligible if any(r in marked_rows for r in u.rows())],
            key=lambda u: (u.length, u.width),
            reverse=True,
        )
        ordered = normal + underfilled

        bins = place_units(
            ordered,
            rule.length_cap,
            rule.max_payload,
            rule.width_hard_limit,
            rule.width_penalty * UNDERFILLED_WIDTH_PENALTY_MULTIPLIER,
            rule.length_fill_weight,
            order_mode="as_is",
        )
        bins = improve_bins(bins, rule.length_cap, rule.max_payload, rule.width_hard_limit)
        partial = bins_to_assignment(bins, f"{code}-")
        out.update(partial)
        assigned = set(partial.keys())
        remaining = [u for u in remaining if all(r not in assigned for r in u.rows())]

    fallback_idx = 1
    fallback_priority = ["20GP", "20FR", "40FR", "40HQ"]
    ordered_fallback = [c for c in fallback_priority if c in allowed_codes]
    for c in allowed_codes:
        if c not in ordered_fallback and c in BOX_RULES:
            ordered_fallback.append(c)

    for u in remaining:
        picked = None
        fr_pick = pick_fr_by_volume_utilization(u, ordered_fallback)
        if fr_pick is not None:
            picked = fr_pick
        for code in ordered_fallback:
            if picked is not None:
                break
            if code in BOARD_CODES and (not unit_can_use_custom_board(u)):
                continue
            if unit_fits_rule(u, BOX_RULES[code]):
                picked = code
                break
        if picked is None and ordered_fallback:
            picked = ordered_fallback[0]
        if picked is None:
            picked = "40FR"
        for r in u.rows():
            out[r] = f"{picked}-FALLBACK{fallback_idx}"
        fallback_idx += 1
    return out


def optimize_with_optional_boxes(units: List[Unit], rules: List[str]) -> Dict[int, str]:
    base = pack_by_rule_priority_mode(units, rules)
    best = global_backoff_optimize(units, rules, base)
    best_obj = assignment_objective(best, units)

    # 20FR retention rule:
    # 1) Compare box count first.
    # 2) When box count ties, compare volume utilization.
    # Keep 20FR when it has higher utilization under equal box count.
    if "20FR" in rules and "40FR" in rules and assignment_uses_model(best, "20FR"):
        rules_no_20fr = [r for r in rules if r != "20FR"]
        base_no = pack_by_rule_priority_mode(units, rules_no_20fr)
        cand_no = global_backoff_optimize(units, rules_no_20fr, base_no)
        obj_no = assignment_objective(cand_no, units)
        no20fr_better = (obj_no[0] < best_obj[0]) or (obj_no[0] == best_obj[0] and obj_no[2] < best_obj[2])
        if no20fr_better:
            best, best_obj = cand_no, obj_no

    # 20GP is only kept when it reduces total box count vs path without 20GP.
    if "20GP" in rules and assignment_uses_model(best, "20GP"):
        rules_no_20gp = [r for r in rules if r != "20GP"]
        base_no = pack_by_rule_priority_mode(units, rules_no_20gp)
        cand_no = global_backoff_optimize(units, rules_no_20gp, base_no)
        obj_no = assignment_objective(cand_no, units)
        if obj_no[0] <= best_obj[0]:
            best, best_obj = cand_no, obj_no

    # Re-pack iteratively when boxes are underfilled by length.
    # In these rounds we bias strongly toward width-similar merging.
    for _ in range(UNDERFILLED_REPACK_MAX_ROUNDS):
        if not has_underfilled_boxes(best, units):
            break
        cand_merge = repack_prioritize_merging_underfilled(units, rules, best)
        cand_merge = consolidate_tiny_underfilled_boxes(cand_merge, units)
        cand_merge = global_backoff_optimize(units, rules, cand_merge)
        obj_merge = assignment_objective(cand_merge, units)
        if obj_merge < best_obj:
            best, best_obj = cand_merge, obj_merge
        else:
            break

    return best


def box_volume_capacity(model: str) -> float:
    rule = BOX_RULES.get(model)
    if rule is None:
        return 0.0
    if model == "20FR":
        eff_width = 2226.0
    elif model == "40FR":
        eff_width = 2374.0
    else:
        eff_width = rule.max_width if rule.max_width is not None else rule.width_hard_limit
    max_h = rule.max_height if rule.max_height is not None else 4500.0
    return max(0.0, rule.length_cap) * max(0.0, eff_width) * max(0.0, max_h)


def assignment_objective(assignments: Dict[int, str], units: List[Unit]) -> Tuple[int, float, float, float]:
    if not assignments:
        return (10**9, 10**9, 10**9, 10**9)
    box_units = box_units_from_assignments(assignments, units)
    if not box_units or len(box_units) == 0:
        return (10**9, 10**9, 10**9, 10**9)
    assigned_unit_count = sum(len(us) for us in box_units.values())
    if assigned_unit_count < len(units):
        return (10**9, 10**9, 10**9, 10**9)

    total_leftover = 0.0
    total_balance_excess = 0.0
    total_cargo_volume = 0.0
    total_box_capacity_volume = 0.0
    for b, us in box_units.items():
        model = b.split("-", 1)[0]
        rule = BOX_RULES.get(model)
        if rule is None:
            continue
        total_leftover += max(0.0, rule.length_cap - used_length_of_units(us))
        total_cargo_volume += sum(max(0.0, u.length) * max(0.0, u.width) * max(0.0, u.height) for u in us)
        total_box_capacity_volume += box_volume_capacity(model)
        if any(u.weight >= BIG_PIECE_WEIGHT for u in us):
            total_balance_excess += max(0.0, estimate_side_diff(us) - SIDE_WEIGHT_DIFF_LIMIT)
    utilization_penalty = (total_box_capacity_volume / total_cargo_volume) if total_cargo_volume > 0 else 10**9
    # Objective order:
    # 1) minimize box count
    # 2) minimize side-balance excess
    # 3) maximize volume utilization (implemented as minimizing capacity/cargo ratio)
    # 4) minimize leftover length
    return (len(box_units), total_balance_excess, utilization_penalty, total_leftover)


def global_backoff_optimize(units: List[Unit], rule_codes: List[str], base: Dict[int, str]) -> Dict[int, str]:
    best = base.copy()
    best_obj = assignment_objective(best, units)
    strategies: List[Tuple[str, int]] = [("length", 0), ("weight", 0), ("hybrid", 0), ("width", 0)]
    strategies.extend([("random", i) for i in range(1, 9)])
    for mode, seed in strategies:
        cand = pack_by_rule_priority_mode(units, rule_codes, order_mode=mode, random_seed=seed)
        obj = assignment_objective(cand, units)
        if obj < best_obj:
            best = cand
            best_obj = obj
    return best


def pack_hq(units: List[Unit]) -> Dict[int, str]:
    rule = BOX_RULES["40HQ"]
    bins = place_units(units, rule.length_cap, rule.max_payload, 900, 2.0, 1.0)
    bins = improve_bins(bins, rule.length_cap, rule.max_payload, 900)
    return bins_to_assignment(bins, "40HQ-")


def build_parallel_merge_rows(units: List[Unit]) -> List[List[int]]:
    merges: List[List[int]] = []
    for u in units:
        if len(u.items) <= 1:
            continue
        rows: List[int] = []
        for it in u.items:
            rows.extend(it.bound_rows if it.bound_rows else [it.row])
        merges.append(sorted(rows))
    return merges


def pack_scenario1(items: List[Item]) -> Tuple[Dict[int, str], List[List[int]]]:
    hq_units, fr_items = build_parallel_groups(items)
    hq_rule = BOX_RULES["40HQ"]
    hq_usable = [u for u in hq_units if unit_fits_rule(u, hq_rule)]
    hq_reject = [it for u in hq_units if not unit_fits_rule(u, hq_rule) for it in u.items]

    # initial HQ backoff
    base_items = fr_items + hq_reject
    idx_all = list(range(len(hq_usable)))

    def eval_selected(sel: List[int]) -> Tuple[int, int, Dict[int, str], List[Unit]]:
        selected = [hq_usable[i] for i in sel]
        moved = [it for i, u in enumerate(hq_usable) if i not in set(sel) for it in u.items]
        ass_hq = pack_hq(selected) if selected else {}
        ass_fr = pack_by_rule_priority_mode([Unit.from_item(i) for i in (base_items + moved)], ["40FR", "20FR", *BOARD_CODES])
        ass = {**ass_fr, **ass_hq}
        return len(set(ass.values())), len({v for v in ass_hq.values() if str(v).startswith("40HQ-")}), ass, selected

    best_sel = idx_all[:]
    best_cnt, best_hq, best_ass, best_units = eval_selected(best_sel)
    improved = True
    while improved and best_sel:
        improved = False
        cand = None
        for i in best_sel:
            trial = [x for x in best_sel if x != i]
            cnt, hq_cnt, ass, sel_units = eval_selected(trial)
            key = (cnt, hq_cnt)
            if key < (best_cnt, best_hq):
                if cand is None or key < cand[0]:
                    cand = (key, trial, ass, sel_units)
        if cand is not None:
            (best_cnt, best_hq), best_sel, best_ass, best_units = cand
            improved = True

    # unified global backoff (must include moved-out HQ-eligible items as FR singles)
    moved_out_items = [it for i, u in enumerate(hq_usable) if i not in set(best_sel) for it in u.items]
    final_units = best_units + [Unit.from_item(i) for i in (base_items + moved_out_items)]
    final_ass = global_backoff_optimize(final_units, ["40HQ", "40FR", "20FR", *BOARD_CODES], best_ass)
    # Keep original return shape; final parallel merge/repack is applied globally in generate_outputs.
    merges = build_parallel_merge_rows(best_units)
    return final_ass, merges


def choose_force_box_for_item(item: Item, allowed_codes: List[str] | None = None) -> str:
    unit = Unit.from_item(item)
    allowed = set(allowed_codes or BOX_RULES.keys())
    for code in ["20GP", "40HQ"]:
        if code in allowed and unit_fits_rule(unit, BOX_RULES[code]):
            return code
    fr_pick = pick_fr_by_volume_utilization(unit, [c for c in ("20FR", "40FR") if c in allowed])
    if fr_pick is not None:
        return fr_pick
    if "710板" in allowed and unit_can_use_custom_board(unit):
        return "710板"
    for code in ["40FR", "20FR", "20GP", "40HQ"]:
        if code in allowed:
            return code
    return "40FR"



def enforce_all_items_assigned(
    assignments: Dict[int, str], items: List[Item], allowed_codes: List[str] | None = None
) -> Dict[int, str]:
    out = assignments.copy()
    force_seq: Dict[str, int] = {}
    for item in items:
        rows = item.bound_rows if item.bound_rows else [item.row]
        if all(r in out for r in rows):
            continue
        model = choose_force_box_for_item(item, allowed_codes=allowed_codes)
        force_seq[model] = force_seq.get(model, 0) + 1
        box = f"{model}-FORCE{force_seq[model]}"
        for r in rows:
            out[r] = box
    return out


def optimize_leftover_box_model(
    assignments: Dict[int, str], items: List[Item], allowed_codes: List[str] | None = None
) -> Dict[int, str]:
    """
    For leftover-like single-item boxes:
    keep 20GP preference first, then choose 20FR/40FR by higher
    volume utilization ratio when both are feasible.
    """
    out = assignments.copy()
    allowed = set(allowed_codes or BOX_RULES.keys())
    row_to_item: Dict[int, Item] = {}
    for item in items:
        for r in (item.bound_rows if item.bound_rows else [item.row]):
            row_to_item[r] = item

    box_to_items: Dict[str, Dict[int, Item]] = {}
    for r, box in out.items():
        it = row_to_item.get(r)
        if it is None:
            continue
        key = str(box)
        box_to_items.setdefault(key, {})
        box_to_items[key][it.row] = it

    rename_map: Dict[str, str] = {}
    for box_name, item_map in box_to_items.items():
        if len(item_map) != 1:
            continue
        model = box_name.split("-", 1)[0]
        if model not in {"40HQ", "40FR", "20FR"}:
            continue
        item = next(iter(item_map.values()))
        unit = Unit.from_item(item)
        if "20GP" in allowed and unit_fits_rule(unit, BOX_RULES["20GP"]):
            target = "20GP"
        else:
            fr_pick = pick_fr_by_volume_utilization(unit, [c for c in ("20FR", "40FR") if c in allowed])
            if fr_pick is not None:
                target = fr_pick
            elif "40FR" in allowed and model in {"40HQ", "20FR"} and unit_fits_rule(unit, BOX_RULES["40FR"]):
                target = "40FR"
            else:
                continue
        suffix = box_name.split("-", 1)[1] if "-" in box_name else "1"
        rename_map[box_name] = f"{target}-{suffix}"

    # Extra rule: downgrade small-occupancy 40FR boxes to 20FR when feasible.
    # This addresses very low length-utilization 40FR boxes containing a few small pieces.
    for box_name, item_map in box_to_items.items():
        if box_name in rename_map:
            continue
        model = box_name.split("-", 1)[0]
        if model != "40FR":
            continue
        if "20FR" not in allowed:
            continue
        if len(item_map) <= 1:
            continue
        items_in_box = list(item_map.values())
        total_len = sum(i.length for i in items_in_box)
        total_wt = sum(i.weight for i in items_in_box)
        # Keep a practical downgrade guard: only for clearly underfilled 40FR.
        if total_len > BOX_RULES["40FR"].length_cap * 0.45:
            continue
        if total_len > BOX_RULES["20FR"].length_cap:
            continue
        if BOX_RULES["20FR"].max_payload is not None and total_wt > BOX_RULES["20FR"].max_payload:
            continue
        if any(not unit_fits_rule(Unit.from_item(it), BOX_RULES["20FR"]) for it in items_in_box):
            continue
        suffix = box_name.split("-", 1)[1] if "-" in box_name else "1"
        rename_map[box_name] = f"20FR-{suffix}"

    if not rename_map:
        return out
    for r, box in list(out.items()):
        b = str(box)
        if b in rename_map:
            out[r] = rename_map[b]
    return out


def pack_scenario2(items: List[Item]) -> Dict[int, str]:
    units = [Unit.from_item(i) for i in items]
    return optimize_with_optional_boxes(units, ["40FR", "20FR", *BOARD_CODES])


def pack_scenario3(items: List[Item]) -> Dict[int, str]:
    units = [Unit.from_item(i) for i in items]
    rules = ["40HQ", "40FR", "20FR", "40GP", "20GP", *BOARD_CODES]
    return optimize_with_optional_boxes(units, rules)

def pack_custom(items: List[Item], box_codes: List[str]) -> Dict[int, str]:
    valid = [c for c in box_codes if c in BOX_RULES]
    if not valid:
        raise ValueError("??????????????????")
    uniq = list(dict.fromkeys(valid))
    # Keep user-defined priority order as-is.
    for b in BOARD_CODES:
        if b not in uniq:
            uniq.append(b)
    units = [Unit.from_item(i) for i in items]
    return optimize_with_optional_boxes(units, uniq)


def pack_auto(items: List[Item], use_hq: bool = False) -> Dict[int, str]:
    units = [Unit.from_item(i) for i in items]
    # Auto mode disables 40GP from the decision flow by default.
    rules = ["40FR", "20FR", "20GP", *BOARD_CODES]
    if use_hq:
        rules = ["40HQ", "40FR", "20FR", "20GP", *BOARD_CODES]
    return optimize_with_optional_boxes(units, rules)


def clear_merges(ws):
    # Clear output-area merges including formula block (M~X), to avoid stale merge conflicts.
    ranges = [r for r in ws.merged_cells.ranges if r.min_row >= DATA_START_ROW and r.min_col <= 24]
    for r in ranges:
        ws.unmerge_cells(str(r))


def apply_output_style(ws, start_row: int, end_row: int, end_col: int = REMARK_COL):
    for r in range(start_row, end_row + 1):
        for c in range(1, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.font = OUTPUT_FONT
            cell.alignment = OUTPUT_ALIGN
            cell.border = OUTPUT_BORDER


def order_boxes(assignments: Dict[int, str], merge_groups: List[List[int]]) -> List[Tuple[str, List[int]]]:
    grouped: Dict[str, List[int]] = {}
    for row, box in assignments.items():
        grouped.setdefault(str(box), []).append(row)
    anchor: Dict[int, int] = {}
    for g in merge_groups:
        if not g:
            continue
        a = min(g)
        for r in g:
            anchor[r] = a
    rank = {"20GP": 1, "40GP": 2, "40HQ": 3, "40FR": 4, "20FR": 5}
    for i, b in enumerate(BOARD_CODES, start=6):
        rank[b] = i
    names = sorted(grouped.keys(), key=lambda b: (rank.get(b.split("-", 1)[0], 99), min(grouped[b])))
    out: List[Tuple[str, List[int]]] = []
    for n in names:
        rs = grouped[n]
        rs.sort(key=lambda r: (anchor.get(r, r), r))
        out.append((n, rs))
    return out


def merge_car_info_runs(ws, start_row: int, end_row: int):
    run_start = start_row
    run_val = ws.cell(row=start_row, column=CAR_INFO_SEQ_COL).value
    for row in range(start_row + 1, end_row + 2):
        v = ws.cell(row=row, column=CAR_INFO_SEQ_COL).value if row <= end_row else None
        if v != run_val:
            if run_val not in (None, "") and row - 1 > run_start:
                ws.merge_cells(start_row=run_start, start_column=CAR_INFO_SEQ_COL, end_row=row - 1, end_column=CAR_INFO_SEQ_COL)
                ws.merge_cells(start_row=run_start, start_column=CAR_INFO_COL, end_row=row - 1, end_column=CAR_INFO_COL)
            run_start = row
            run_val = v


def merge_box_formula_runs(ws, start_row: int, end_row: int, start_col: int = 13, end_col: int = 23):
    """
    Merge per-box calculated columns (M~X) by contiguous Car Info Seq runs.
    """
    run_start = start_row
    run_val = ws.cell(row=start_row, column=CAR_INFO_SEQ_COL).value
    for row in range(start_row + 1, end_row + 2):
        v = ws.cell(row=row, column=CAR_INFO_SEQ_COL).value if row <= end_row else None
        if v != run_val:
            if run_val not in (None, "") and row - 1 > run_start:
                for col in range(start_col, end_col + 1):
                    ws.merge_cells(start_row=run_start, start_column=col, end_row=row - 1, end_column=col)
            run_start = row
            run_val = v


def append_row_remark(ws, row: int, text: str):
    cur = ws.cell(row=row, column=REMARK_COL).value
    cur_s = str(cur).strip() if cur is not None else ""
    parts = [p for p in cur_s.replace("；", ";").split(";") if p]
    if text not in parts:
        parts.append(text)
    ws.cell(row=row, column=REMARK_COL, value="；".join(parts))


def build_over_limit_remark(model: str, width: float | None, height: float | None) -> str | None:
    if model not in FR_WIDTH_REMARK_LIMITS:
        return None
    over_w = (width is not None) and (width > FR_WIDTH_REMARK_LIMITS[model])
    over_h = (height is not None) and (height > FR_HEIGHT_REMARK_LIMIT)
    if over_w and over_h:
        return "双超"
    if over_w:
        return "超宽"
    if over_h:
        return "超高"
    return None


def fill_box_formula_cells(
    ws,
    cargo_rows: List[int],
    box_seq_by_row: Dict[int, int],
    summary_styles: Dict[int, object] | None = None,
    formula_styles: Dict[int, object] | None = None,
):
    # Top summary formulas.
    ws.cell(row=3, column=9, value="=SUM(V6:V505)")   # I3
    ws.cell(row=3, column=11, value="=G3-I3")         # K3
    ws.cell(row=3, column=13, value="=SUM(U6:U1005)") # M3
    for c in (9, 11, 13):
        cell = ws.cell(row=3, column=c)
        if summary_styles and c in summary_styles and summary_styles[c] is not None:
            cell._style = copy(summary_styles[c])

    for r in cargo_rows:
        seq_no = box_seq_by_row.get(r)
        if seq_no is None:
            continue
        # X: formula sequence id; same box -> same id.
        ws.cell(row=r, column=24, value=seq_no)

        # M~W formula block (match template logic).
        ws.cell(row=r, column=13, value=f"=SUMIF($X$6:$X$1000,X{r},$E$6:$E$1000)")  # M
        ws.cell(row=r, column=14, value=f"=MAXIFS($F$6:$F$1000,$X$6:$X$1000,X{r})") # N
        ws.cell(row=r, column=15, value=f"=MAXIFS($G$6:$G$1000,$X$6:$X$1000,X{r})") # O
        ws.cell(row=r, column=16, value=f"=SUMIF($X$6:$X$1000,X{r},$I$6:$I$1000)")   # P
        ws.cell(row=r, column=17, value=f"=SUMIF($X$6:$X$1000,X{r},$H$6:$H$1000)")   # Q
        ws.cell(
            row=r,
            column=18,
            value=(
                f'=IF(J{r}="40FR",12064,'
                f'IF(J{r}="20FR",6038,'
                f'IF(J{r}="40GP",12192,'
                f'IF(J{r}="40HQ",12192,'
                f'IF(OR(J{r}="710板",J{r}="710定制板"),6700,'
                f'IF(OR(J{r}="880板",J{r}="880定制板"),8100,0))))))'
            ),
        )  # R
        ws.cell(
            row=r,
            column=19,
            value=(
                f'=IF(J{r}="40GP",2438,'
                f'IF(J{r}="40HQ",2438,'
                f'IF(OR(J{r}="40FR",J{r}="20FR"),'
                f'MAX(2374,MAXIFS($F$6:$F$1000,$X$6:$X$1000,X{r})),'
                f'IF(OR(J{r}="710板",J{r}="710定制板",J{r}="880板",J{r}="880定制板"),'
                f'MAX(5800,MAXIFS($F$6:$F$1000,$X$6:$X$1000,X{r})),'
                f'MAXIFS($F$6:$F$1000,$X$6:$X$1000,X{r})))))'
            ),
        )  # S
        ws.cell(
            row=r,
            column=20,
            value=(
                f'=IF(J{r}="40FR",MAX(MAXIFS($G$6:$G$1000,$X$6:$X$1000,X{r})+648,1900),'
                f'IF(J{r}="20FR",MAX(MAXIFS($G$6:$G$1000,$X$6:$X$1000,X{r})+350,1900),'
                f'IF(OR(J{r}="710板",J{r}="710定制板"),MAX(MAXIFS($G$6:$G$1000,$X$6:$X$1000,X{r})+400,1900),'
                f'IF(OR(J{r}="880板",J{r}="880定制板"),MAX(MAXIFS($G$6:$G$1000,$X$6:$X$1000,X{r})+400,1900),'
                f'IF(J{r}="40GP",2896,IF(J{r}="40HQ",2896,0))))))'
            ),
        )  # T
        ws.cell(row=r, column=21, value=f"=R{r}*S{r}/1000000")      # U
        ws.cell(row=r, column=22, value=f"=R{r}*S{r}*T{r}/1000000000")  # V
        ws.cell(
            row=r,
            column=23,
            value=(
                f'=IF(J{r}="40FR",SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+5100,'
                f'IF(J{r}="20FR",SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+2800,'
                f'IF(J{r}="40GP",SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+4000,'
                f'IF(J{r}="40HQ",SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+3700,'
                f'IF(OR(J{r}="710板",J{r}="710定制板"),SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+0,'
                f'IF(OR(J{r}="880板",J{r}="880定制板"),SUMIF($X$6:$X$1005,X{r},$H$6:$H$1005)+0,0))))))'
            ),
        )  # W

        for c in range(13, 25):
            cell = ws.cell(row=r, column=c)
            if formula_styles and c in formula_styles and formula_styles[c] is not None:
                cell._style = copy(formula_styles[c])


def apply_uniform_generated_styles(ws, rows: List[int], style_map: Dict[int, object]):
    if not rows:
        return
    for r in sorted(set(rows)):
        for c, st in style_map.items():
            if st is None:
                continue
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell._style = copy(st)


def apply_assignments(template_path: Path, output_path: Path, assignments: Dict[int, str], merge_groups: List[List[int]] | None = None):
    wb = load_workbook(template_path)
    ws = wb.active
    summary_style_map = {c: copy(ws.cell(row=3, column=c)._style) for c in (9, 11, 13)}
    generated_style_map = {c: copy(ws.cell(row=DATA_START_ROW, column=c)._style) for c in range(10, 25)}
    formula_style_map = {c: generated_style_map[c] for c in range(13, 25)}

    preserve_cols = {1, 2, 5, 6, 7, 8, 9}
    preserve_merges: List[Tuple[int, int, int]] = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= DATA_START_ROW and mr.min_col == mr.max_col and mr.min_col in preserve_cols:
            preserve_merges.append((mr.min_col, mr.min_row, mr.max_row))

    clear_merges(ws)
    cargo_rows = sorted(assignments.keys())
    if not cargo_rows:
        wb.save(output_path)
        return

    values = {r: [ws.cell(row=r, column=c).value for c in range(1, 10)] for r in range(DATA_START_ROW, ws.max_row + 1)}
    style_map = {
        r: [copy(ws.cell(row=r, column=c)._style) for c in range(1, REMARK_COL + 1)]
        for r in range(DATA_START_ROW, ws.max_row + 1)
    }
    source_active_rows = [
        r for r in range(DATA_START_ROW, ws.max_row + 1) if any(ws.cell(row=r, column=c).value not in (None, "") for c in range(1, 10))
    ]

    ordered_boxes = order_boxes(assignments, merge_groups or [])
    ordered_rows: List[int] = []
    for _, rs in ordered_boxes:
        ordered_rows.extend(rs)
    non_cargo_rows = [r for r in source_active_rows if r not in set(ordered_rows)]
    final_rows = ordered_rows + non_cargo_rows

    for row in range(DATA_START_ROW, max(ws.max_row, DATA_END_ROW) + 1):
        for c in range(1, REMARK_COL + 1):
            ws.cell(row=row, column=c).value = None

    old_to_new: Dict[int, int] = {}
    for idx, old in enumerate(final_rows):
        nr = DATA_START_ROW + idx
        old_to_new[old] = nr
        for c, v in enumerate(values[old], start=1):
            ws.cell(row=nr, column=c, value=v)
        if old in style_map:
            for c in range(1, REMARK_COL + 1):
                ws.cell(row=nr, column=c)._style = copy(style_map[old][c - 1])

    seq_by_model: Dict[str, int] = {}
    global_box_seq: Dict[str, int] = {}
    cargo_new_rows: List[int] = []
    box_seq_by_new_row: Dict[int, int] = {}
    global_seq_counter = 0
    for box_name, rs in ordered_boxes:
        model = str(box_name).split("-", 1)[0]
        display_model = MODEL_DISPLAY_NAMES.get(model, model)
        seq_by_model[model] = seq_by_model.get(model, 0) + 1
        seq = f"{display_model}-{seq_by_model[model]}"
        global_seq_counter += 1
        global_box_seq[box_name] = global_seq_counter
        for old in rs:
            nr = old_to_new[old]
            ws.cell(row=nr, column=CAR_INFO_COL, value=display_model)
            ws.cell(row=nr, column=CAR_INFO_SEQ_COL, value=seq)
            cargo_new_rows.append(nr)
            box_seq_by_new_row[nr] = global_box_seq[box_name]

    fill_box_formula_cells(ws, cargo_new_rows, box_seq_by_new_row, summary_style_map, formula_style_map)

    if ordered_rows:
        merge_box_formula_runs(ws, DATA_START_ROW, DATA_START_ROW + len(ordered_rows) - 1, 13, 23)
        merge_car_info_runs(ws, DATA_START_ROW, DATA_START_ROW + len(ordered_rows) - 1)

    end_row = DATA_START_ROW + len(final_rows) - 1 if final_rows else DATA_START_ROW

    # FR cargo remark enrichment:
    # - 超宽: compare width against FR model-specific limit.
    # - 超高: height > 1900.
    for row in range(DATA_START_ROW, end_row + 1):
        model = ws.cell(row=row, column=CAR_INFO_COL).value
        if model not in FR_WIDTH_REMARK_LIMITS:
            continue
        width_v = ws.cell(row=row, column=6).value
        height_v = ws.cell(row=row, column=7).value
        try:
            width = float(width_v) if width_v is not None else None
            height = float(height_v) if height_v is not None else None
        except Exception:
            continue
        over_remark = build_over_limit_remark(str(model), width, height)
        if over_remark:
            append_row_remark(ws, row, over_remark)

    if merge_groups:
        for g in merge_groups:
            mapped = sorted(old_to_new[r] for r in g if r in old_to_new)
            if len(mapped) <= 1:
                continue
            labels: List[str] = ["并列"]
            for r in mapped:
                v = ws.cell(row=r, column=REMARK_COL).value
                if v is None:
                    continue
                for p in str(v).replace("；", ";").split(";"):
                    p = p.strip()
                    if p and p not in labels:
                        labels.append(p)
            if all(b - a == 1 for a, b in zip(mapped, mapped[1:])):
                ws.merge_cells(start_row=mapped[0], start_column=REMARK_COL, end_row=mapped[-1], end_column=REMARK_COL)
                ws.cell(row=mapped[0], column=REMARK_COL, value="；".join(labels))
            else:
                text = "；".join(labels)
                for r in mapped:
                    ws.cell(row=r, column=REMARK_COL, value=text)

    for col, s, e in preserve_merges:
        source_rows = [r for r in range(s, e + 1) if r in old_to_new]
        if len(source_rows) <= 1:
            continue
        mapped = sorted(old_to_new[r] for r in source_rows)
        if all(b - a == 1 for a, b in zip(mapped, mapped[1:])):
            ws.merge_cells(start_row=mapped[0], start_column=col, end_row=mapped[-1], end_column=col)

    # Final pass: normalize all generated-content styles to template baseline.
    apply_uniform_generated_styles(ws, cargo_new_rows, generated_style_map)
    wb.save(output_path)


def generate_outputs(
    input_path: Path,
    outdir: Path,
    scenario: str = "both",
    custom_boxes: List[str] | None = None,
    auto_use_hq: bool = False,
) -> List[Tuple[str, Path]]:
    items = load_items(input_path)
    if not items:
        raise ValueError("No cargo rows detected; please check workbook format.")
    outdir.mkdir(parents=True, exist_ok=True)
    outputs: List[Tuple[str, Path]] = []

    if scenario in ("scenario1", "both", "all"):
        ass, merges = pack_scenario1(items)
        allowed = ["40HQ", "40FR", "20FR", *BOARD_CODES]
        ass = enforce_all_items_assigned(ass, items, allowed_codes=allowed)
        ass = optimize_leftover_box_model(ass, items, allowed_codes=allowed)
        ass, merges = apply_parallel_repack_and_merge(ass, items, allowed_codes=allowed)
        p = outdir / f"{input_path.stem}-scenario1-40HQ+FR.xlsx"
        apply_assignments(input_path, p, ass, merges or None)
        outputs.append((SCENARIO_LABELS["scenario1"], p))
    if scenario in ("scenario2", "both", "all"):
        ass = pack_scenario2(items)
        allowed = ["40FR", "20FR", *BOARD_CODES]
        ass = enforce_all_items_assigned(ass, items, allowed_codes=allowed)
        ass = optimize_leftover_box_model(ass, items, allowed_codes=allowed)
        ass, merges = apply_parallel_repack_and_merge(ass, items, allowed_codes=allowed)
        p = outdir / f"{input_path.stem}-scenario2-FR-only.xlsx"
        apply_assignments(input_path, p, ass, merges or None)
        outputs.append((SCENARIO_LABELS["scenario2"], p))
    if scenario in ("scenario3", "all"):
        ass = pack_scenario3(items)
        allowed = ["40HQ", "40FR", "20FR", "40GP", "20GP", *BOARD_CODES]
        ass = enforce_all_items_assigned(ass, items, allowed_codes=allowed)
        ass = optimize_leftover_box_model(ass, items, allowed_codes=allowed)
        ass, merges = apply_parallel_repack_and_merge(ass, items, allowed_codes=allowed)
        p = outdir / f"{input_path.stem}-scenario3-GP+HQ+FR.xlsx"
        apply_assignments(input_path, p, ass, merges or None)
        outputs.append((SCENARIO_LABELS["scenario3"], p))
    if scenario in ("auto", "all"):
        ass = pack_auto(items, use_hq=auto_use_hq)
        allowed = (["40HQ"] if auto_use_hq else []) + ["40FR", "20FR", "20GP", *BOARD_CODES]
        ass = enforce_all_items_assigned(ass, items, allowed_codes=allowed)
        ass = optimize_leftover_box_model(ass, items, allowed_codes=allowed)
        ass, merges = apply_parallel_repack_and_merge(ass, items, allowed_codes=allowed)
        suffix = "with-HQ" if auto_use_hq else "no-HQ"
        p = outdir / f"{input_path.stem}-auto-{suffix}.xlsx"
        apply_assignments(input_path, p, ass, merges or None)
        outputs.append((SCENARIO_LABELS["auto"], p))
    if scenario == "custom":
        ass = pack_custom(items, custom_boxes or [])
        allowed = [c for c in (custom_boxes or []) if c in BOX_RULES]
        for b in BOARD_CODES:
            if b not in allowed:
                allowed.append(b)
        ass = enforce_all_items_assigned(ass, items, allowed_codes=allowed)
        ass = optimize_leftover_box_model(ass, items, allowed_codes=allowed)
        ass, merges = apply_parallel_repack_and_merge(ass, items, allowed_codes=allowed)
        p = outdir / f"{input_path.stem}-custom-{'-'.join(custom_boxes or [])}.xlsx"
        apply_assignments(input_path, p, ass, merges or None)
        outputs.append((SCENARIO_LABELS["custom"], p))
    return outputs


def main():
    parser = argparse.ArgumentParser(description="Generate packing plans")
    parser.add_argument("input", type=Path)
    parser.add_argument("--outdir", type=Path, default=Path.cwd())
    parser.add_argument(
        "--scenario",
        choices=["scenario1", "scenario2", "scenario3", "auto", "both", "all", "custom"],
        default="both",
    )
    parser.add_argument("--boxes", default="")
    parser.add_argument("--auto-use-hq", action="store_true")
    args = parser.parse_args()

    boxes = [x.strip() for x in args.boxes.split(",") if x.strip()]
    outputs = generate_outputs(args.input, args.outdir, args.scenario, boxes, args.auto_use_hq)
    for name, path in outputs:
        print(f"{name} output: {path}")


if __name__ == "__main__":
    main()
